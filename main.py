import os
import re
import log_to_csv_module1 as m1
import log_to_csv_module2 as m2
import log_to_csv_module3 as m3
import log_to_csv_module4 as m4
import log_to_csv_module4_DT as DT
from openpyxl import load_workbook

listOfFiles = []
path = '/Users/justinclay/Downloads/AKB Log files'
Excelfile = '/Users/justinclay/Downloads/Steam Akribis Data Lot 10-11 to 10-14.xlsx'

m1_finallist = []
m2_finallist = []
m3_finallist = []
m4_finallist = []

print(f'Opening Excel Sheet:{Excelfile}')
wb = load_workbook(filename=Excelfile)
print(f'Sheet Opened:{Excelfile}')

print('Creating Sheets')
sheet_names = wb.sheetnames
sheet_m1 = wb['Summary M1']
sheet_m2 = wb['Summary M2']
sheet_m3 = wb['Summary M3']
sheet_m4 = wb['Summary M4']
sheet_DT = wb['M1 Alarm History']
print('Sheets Complete')


def scandirectory(path):
    global listOfFiles
    for (dirpath, dirnames, filenames) in os.walk(path):
        listOfFiles += [os.path.join(dirpath, file) for file in filenames]


l_values = []
csvmaster = [['Day', '', 'Module', '', 'Start', 'End', 'Failure Mode', 'Time(min)', 'MTBA(min)']]

if __name__ == '__main__':
    print('Scanning Directory...')
    scandirectory(path)
    print('Directory Scan Complete')
    print('Remove .DS_Store')
    for i in listOfFiles:
        if i.split("/")[-1] == '.DS_Store':
            listOfFiles.remove(i)
            pass
    print('Remove .DS_Store Complete')

for i in listOfFiles:

    if re.findall('AKB M1', i):
        print('Parsing:', i)
        try:
            m1csv = DT.main('M1', i)
            for i1 in m1csv:
                csvmaster.append(i1)
            _, _, _, _, _, v, M1_new_all_counts = m1.main(i)
            M1_new_all_counts.update({'Build': ""})
            M1_new_all_counts.move_to_end('Build', last=False)
            M1_new_all_counts.update({'Date': i.split("/")[-1].split("_")[0]})
            M1_new_all_counts.move_to_end('Date', last=False)
            print('All Values Extracted:', i)
            m1_finallist.append(M1_new_all_counts)
            m1_finallist.append(list(M1_new_all_counts.values()))
            sheet_m1.append(list(M1_new_all_counts.values()))
        except Exception as e:
            print(e)

    if re.findall('AKB M2', i):
        print('Parsing:', i)
        try:
            m2csv = DT.main('M2', i)
            for i2 in m2csv:
                csvmaster.append(i2)
            _, _, _, _, _, _, M2_new_all_counts = m2.main(i)
            M2_new_all_counts.update({'Build': ""})
            M2_new_all_counts.move_to_end('Build', last=False)
            M2_new_all_counts.update({'Date': i.split("/")[-1].split("_")[0]})
            M2_new_all_counts.move_to_end('Date', last=False)
            print('All Values Extracted:', i)
            m2_finallist.append(M2_new_all_counts)
            m2_finallist.append(list(M2_new_all_counts.values()))
            sheet_m2.append(list(M2_new_all_counts.values()))
        except Exception as e:
            print(e)
    if re.findall('AKB M3', i):
        print('Parsing:', i)
        try:
            m3csv = DT.main('M3', i)
            for i3 in m3csv:
                csvmaster.append(i3)
            _, _, _, _, _, _, M3_new_all_counts = m3.main(i)
            M3_new_all_counts.update({'Build': ""})
            M3_new_all_counts.move_to_end('Build', last=False)
            M3_new_all_counts.update({'Date': i.split("/")[-1].split("_")[0]})
            M3_new_all_counts.move_to_end('Date', last=False)
            print('All Values Extracted:', i)
            m3_finallist.append(M3_new_all_counts)
            m3_finallist.append(list(M3_new_all_counts.values()))
            sheet_m3.append(list(M3_new_all_counts.values()))
        except Exception as e:
            print(e)
    if re.findall('AKB M4', i):
        print('Parsing:', i)
        try:
            m4csv = DT.main('M4', i)
            for i4 in m4csv:
                csvmaster.append(i4)
            _, _, _, _, _, _, M4_new_all_counts = \
                m4.main(i)
            M4_new_all_counts.update({'Build': ""})
            M4_new_all_counts.move_to_end('Build', last=False)
            M4_new_all_counts.update({'Date': i.split("/")[-1].split("_")[0]})
            M4_new_all_counts.move_to_end('Date', last=False)
            print('All Values Extracted:', i.split("/")[-1].split("_")[0])
            m4_finallist.append(M4_new_all_counts)
            m4_finallist.append(list(M4_new_all_counts.values()))
            sheet_m4.append(list(M4_new_all_counts.values()))
        except Exception as e:
            print(e)
for items in csvmaster:
    sheet_DT.append(items)
wb.save(filename='/Users/justinclay/Downloads/akParseTemp2.xlsx')
print('Finally...Done!')